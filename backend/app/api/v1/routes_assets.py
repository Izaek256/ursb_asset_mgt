"""Asset routes: list, filter, search, and register assets."""

import io
import uuid
from datetime import date, datetime
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.db import get_db
from app.models.asset import Asset, AssetCondition, AssetStatus, AssetType, SourceType
from app.models.audit_log import AuditLog
from app.api.v1.auth import get_current_user, require_roles

router = APIRouter(prefix="/api/v1/assets", tags=["assets"])


# ── Response Schemas ────────────────────────────────────────────────────────────────

class AssetDetailResponse(BaseModel):
    """Response for GET /api/v1/assets/{id} - full asset with all related data."""
    asset_id: str
    asset_name: str
    asset_type: str
    category: str
    serial_number: str
    condition: str
    status: str
    source_type: str
    procurement_ref: str | None
    cost: float
    acquisition_date: str
    supplier: str
    department: str | None
    current_custodian_id: str | None
    is_active: bool
    created_at: str
    updated_at: str
    current_custodian: dict | None
    assignment_history: List[dict]
    maintenance_history: List[dict]
    transfer_history: List[dict]
    disposal_record: dict | None

    class Config:
        from_attributes = True


class AssetOut(BaseModel):
    asset_id: str
    asset_name: str
    asset_type: str
    category: str
    serial_number: str
    condition: str
    status: str
    cost: float
    acquisition_date: str
    supplier: str
    department: str | None
    created_at: str

    class Config:
        from_attributes = True


class AssetCreate(BaseModel):
    name: str
    asset_type: str
    serial_number: str
    condition: str
    cost: float
    department: Optional[str] = None
    acquisition_date: str
    status: str
    category: str
    supplier: str
    source_type: str


class AssetUpdateRequest(BaseModel):
    """Request body for PUT /api/v1/assets/{id}. All fields optional."""
    asset_name: Optional[str] = None
    category: Optional[str] = None
    condition: Optional[str] = None
    status: Optional[str] = None
    department: Optional[str] = None
    current_custodian_id: Optional[int] = None
    supplier: Optional[str] = None
    procurement_ref: Optional[str] = None
    cost: Optional[float] = None
    # Non-updatable fields (silently ignored if sent):
    # - asset_id: system-generated, immutable
    # - serial_number: physically labelled on asset, immutable after registration
    # - asset_type: fundamental classification, immutable
    # - acquisition_date: historical fact, immutable
    # - source_type: procurement record, immutable


# Valid status transitions
VALID_STATUS_TRANSITIONS = {
    "Active": ["In Storage", "Under Maintenance", "Disposed"],
    "In Storage": ["Active", "Disposed"],
    "Under Maintenance": ["Active", "Disposed"],
    "Disposed": [],  # Terminal state
}


# Map display-friendly status labels → model enum
_STATUS_MAP = {
    "Active": AssetStatus.ACTIVE,
    "In Store": AssetStatus.IN_STORAGE,
    "In Storage": AssetStatus.IN_STORAGE,
}


@router.post("", response_model=AssetOut, status_code=status.HTTP_201_CREATED)
def create_asset(
    body: AssetCreate,
    db: Session = Depends(get_db),
    current_user=Depends(require_roles("Asset Manager")),
):
    """Register a new asset. Only Asset Managers may call this endpoint."""

    # ── Validate status ─────────────────────────────────────────────────────────
    if body.status not in _STATUS_MAP:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Invalid status '{body.status}'. Allowed: Active, In Store.",
        )
    asset_status = _STATUS_MAP[body.status]

    # ── Validate asset_type enum ────────────────────────────────────────────────
    try:
        asset_type_enum = AssetType(body.asset_type)
    except ValueError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Invalid asset type '{body.asset_type}'.",
        )

    # ── Validate condition enum ─────────────────────────────────────────────────
    try:
        condition_enum = AssetCondition(body.condition)
    except ValueError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Invalid condition '{body.condition}'.",
        )

    # ── Validate source_type enum ───────────────────────────────────────────────
    try:
        source_type_enum = SourceType(body.source_type)
    except ValueError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Invalid source type '{body.source_type}'.",
        )

    # ── Validate acquisition_date ───────────────────────────────────────────────
    try:
        parsed_acquisition_date = datetime.strptime(body.acquisition_date, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Invalid acquisition_date '{body.acquisition_date}'. Expected format: YYYY-MM-DD.",
        )

    # ── Generate unique asset ID ────────────────────────────────────────────────
    new_asset_id = f"URSB-{str(uuid.uuid4()).upper()[:8]}"

    # ── Create asset record ─────────────────────────────────────────────────────
    new_asset = Asset(
        asset_id=new_asset_id,
        asset_name=body.name,
        asset_type=asset_type_enum,
        category=body.category,
        serial_number=body.serial_number,
        condition=condition_enum,
        status=asset_status,
        source_type=source_type_enum,
        cost=body.cost,
        acquisition_date=parsed_acquisition_date,
        supplier=body.supplier,
        department=body.department,
    )
    db.add(new_asset)

    # ── Audit log ───────────────────────────────────────────────────────────────
    audit_entry = AuditLog(
        user_id=current_user.user_id,
        action="ASSET_REGISTRATION",
        table_affected="assets",
        record_id=new_asset_id,
        details=f"Asset {body.name} registered with ID {new_asset_id}",
    )
    db.add(audit_entry)
    db.commit()
    db.refresh(new_asset)

    return AssetOut(
        asset_id=new_asset.asset_id,
        asset_name=new_asset.asset_name,
        asset_type=new_asset.asset_type.value if hasattr(new_asset.asset_type, "value") else str(new_asset.asset_type),
        category=new_asset.category,
        serial_number=new_asset.serial_number,
        condition=new_asset.condition.value if hasattr(new_asset.condition, "value") else str(new_asset.condition),
        status=new_asset.status.value if hasattr(new_asset.status, "value") else str(new_asset.status),
        cost=float(new_asset.cost),
        acquisition_date=str(new_asset.acquisition_date),
        supplier=new_asset.supplier,
        department=new_asset.department,
        created_at=str(new_asset.created_at),
    )


@router.get("", response_model=List[AssetOut])
def list_assets(
    status: Optional[str] = Query(None),
    asset_type: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    _user=Depends(get_current_user),
):
    q = db.query(Asset)
    if status:
        # Convert string to enum - Asset.status is an AssetStatus enum, not a raw string
        try:
            status_enum = AssetStatus(status)
            q = q.filter(Asset.status == status_enum)
        except ValueError:
            valid_values = [e.value for e in AssetStatus]
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Invalid status value. Valid values are: {', '.join(valid_values)}"
            )
    if asset_type:
        # Convert string to enum - Asset.asset_type is an AssetType enum, not a raw string
        try:
            asset_type_enum = AssetType(asset_type)
            q = q.filter(Asset.asset_type == asset_type_enum)
        except ValueError:
            valid_values = [e.value for e in AssetType]
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Invalid asset_type value. Valid values are: {', '.join(valid_values)}"
            )
    if search:
        q = q.filter(
            Asset.asset_name.ilike(f"%{search}%")
            | Asset.serial_number.ilike(f"%{search}%")
        )
    assets = q.order_by(Asset.created_at.desc()).all()
    return [
        AssetOut(
            asset_id=a.asset_id,
            asset_name=a.asset_name,
            asset_type=a.asset_type.value if hasattr(a.asset_type, "value") else str(a.asset_type),
            category=a.category,
            serial_number=a.serial_number,
            condition=a.condition.value if hasattr(a.condition, "value") else str(a.condition),
            status=a.status.value if hasattr(a.status, "value") else str(a.status),
            cost=float(a.cost),
            acquisition_date=str(a.acquisition_date),
            supplier=a.supplier,
            department=a.department,
            created_at=str(a.created_at),
        )
        for a in assets
    ]


@router.get("/{asset_id}", response_model=AssetDetailResponse)
def get_asset_detail(
    asset_id: str,
    db: Session = Depends(get_db),
    _user=Depends(get_current_user),
):
    """
    Fetch a single asset by ID with all related data.
    Access: All authenticated roles.
    Returns 404 if asset not found.
    """
    from app.models.assignment import Assignment
    from app.models.maintenance_record import MaintenanceRecord
    from app.models.transfer import Transfer
    from app.models.disposal_record import DisposalRecord

    asset = db.query(Asset).filter(Asset.asset_id == asset_id).first()
    if not asset:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Asset with ID {asset_id} not found"
        )

    # Current custodian
    current_custodian = None
    if asset.current_custodian:
        current_custodian = {
            "id": str(asset.current_custodian.id),
            "first_name": asset.current_custodian.first_name,
            "last_name": asset.current_custodian.last_name,
            "email": asset.current_custodian.email,
            "department": asset.current_custodian.department,
        }

    # Assignment history
    assignment_history = []
    for assignment in sorted(asset.assignments, key=lambda x: x.assignment_date, reverse=True):
        assignment_history.append({
            "assignment_id": assignment.assignment_id,
            "assigned_to_name": assignment.assigned_to_user.full_name if assignment.assigned_to_user else None,
            "assigned_by_name": assignment.assigned_by_user.full_name if assignment.assigned_by_user else None,
            "assignment_date": str(assignment.assignment_date),
            "return_date": str(assignment.return_date) if assignment.return_date else None,
            "status": assignment.status.value if hasattr(assignment.status, "value") else str(assignment.status),
            "notes": assignment.notes,
        })

    # Maintenance history
    maintenance_history = []
    for maintenance in sorted(asset.maintenance_records, key=lambda x: x.service_date, reverse=True):
        maintenance_history.append({
            "maintenance_id": maintenance.maintenance_id,
            "service_date": str(maintenance.service_date),
            "service_provider": maintenance.service_provider,
            "description": maintenance.description,
            "cost": float(maintenance.cost),
            "next_service_date": str(maintenance.next_service_date) if maintenance.next_service_date else None,
        })

    # Transfer history
    transfer_history = []
    for transfer in sorted(asset.transfers, key=lambda x: x.transfer_date, reverse=True):
        transfer_history.append({
            "transfer_id": transfer.transfer_id,
            "from_user_name": transfer.from_user.full_name if transfer.from_user else None,
            "to_user_name": transfer.to_user.full_name if transfer.to_user else None,
            "transfer_date": str(transfer.transfer_date),
            "reason": transfer.reason,
            "acknowledged_at": str(transfer.acknowledged_at) if transfer.acknowledged_at else None,
        })

    # Disposal record
    disposal_record = None
    if asset.disposal_records:
        disposal = asset.disposal_records[0]  # Should only be one
        disposal_record = {
            "disposal_date": str(disposal.disposal_date),
            "disposal_method": disposal.disposal_method.value if hasattr(disposal.disposal_method, "value") else str(disposal.disposal_method),
            "reason": disposal.reason,
            "authorised_by_name": disposal.authorised_by_user.full_name if disposal.authorised_by_user else None,
        }

    return AssetDetailResponse(
        asset_id=asset.asset_id,
        asset_name=asset.asset_name,
        asset_type=asset.asset_type.value if hasattr(asset.asset_type, "value") else str(asset.asset_type),
        category=asset.category,
        serial_number=asset.serial_number,
        condition=asset.condition.value if hasattr(asset.condition, "value") else str(asset.condition),
        status=asset.status.value if hasattr(asset.status, "value") else str(asset.status),
        source_type=asset.source_type.value if hasattr(asset.source_type, "value") else str(asset.source_type),
        procurement_ref=asset.procurement_ref,
        cost=float(asset.cost),
        acquisition_date=str(asset.acquisition_date),
        supplier=asset.supplier,
        department=asset.department,
        current_custodian_id=str(asset.current_custodian_id) if asset.current_custodian_id else None,
        is_active=asset.is_active,
        created_at=str(asset.created_at),
        updated_at=str(asset.updated_at),
        current_custodian=current_custodian,
        assignment_history=assignment_history,
        maintenance_history=maintenance_history,
        transfer_history=transfer_history,
        disposal_record=disposal_record,
    )


@router.put("/{asset_id}", response_model=AssetDetailResponse)
def update_asset(
    asset_id: str,
    body: AssetUpdateRequest,
    db: Session = Depends(get_db),
    current_user=Depends(require_roles("Asset Manager", "System Administrator")),
):
    """
    Update an asset. Only Asset Manager and System Administrator may call this endpoint.
    Validates status transitions and checks if asset is active and not disposed.
    """
    asset = db.query(Asset).filter(Asset.asset_id == asset_id).first()
    if not asset:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Asset with ID {asset_id} not found"
        )

    # Check if asset is disposed (terminal state)
    if asset.status == AssetStatus.DISPOSED:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Cannot update a disposed asset"
        )

    # Check if asset is inactive
    if not asset.is_active:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Cannot update an inactive asset"
        )

    # Validate status transition if status is being updated
    if body.status is not None:
        current_status = asset.status.value if hasattr(asset.status, "value") else str(asset.status)
        allowed_transitions = VALID_STATUS_TRANSITIONS.get(current_status, [])
        if body.status not in allowed_transitions:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Invalid status transition from {current_status} to {body.status}. Allowed: {', '.join(allowed_transitions) or 'none'}"
            )

    # Update only provided fields
    if body.asset_name is not None:
        asset.asset_name = body.asset_name
    if body.category is not None:
        asset.category = body.category
    if body.condition is not None:
        try:
            asset.condition = AssetCondition(body.condition)
        except ValueError:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Invalid condition '{body.condition}'"
            )
    if body.status is not None:
        try:
            asset.status = AssetStatus(body.status)
        except ValueError:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Invalid status '{body.status}'"
            )
    if body.department is not None:
        asset.department = body.department
    if body.current_custodian_id is not None:
        asset.current_custodian_id = str(body.current_custodian_id)
    if body.supplier is not None:
        asset.supplier = body.supplier
    if body.procurement_ref is not None:
        asset.procurement_ref = body.procurement_ref
    if body.cost is not None:
        asset.cost = body.cost

    # Audit log
    audit_entry = AuditLog(
        user_id=current_user.user_id,
        action="ASSET_UPDATE",
        table_affected="assets",
        record_id=asset_id,
        details=f"Asset {asset.asset_name} updated",
    )
    db.add(audit_entry)
    db.commit()
    db.refresh(asset)

    # Return updated asset detail by calling the get function
    return get_asset_detail(asset_id, db, current_user)


@router.patch("/{asset_id}/deactivate")
def deactivate_asset(
    asset_id: str,
    db: Session = Depends(get_db),
    current_user=Depends(require_roles("Asset Manager", "System Administrator")),
):
    """
    Deactivate an asset. Sets is_active = False.
    Requires: asset not inactive, no active assignment, not Disposed.
    """
    from app.models.assignment import Assignment, AssignmentStatus

    asset = db.query(Asset).filter(Asset.asset_id == asset_id).first()
    if not asset:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Asset with ID {asset_id} not found"
        )

    # Check if already inactive
    if not asset.is_active:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Asset is already inactive"
        )

    # Check if disposed
    if asset.status == AssetStatus.DISPOSED:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Cannot deactivate a disposed asset"
        )

    # Check for active assignment
    active_assignment = db.query(Assignment).filter(
        Assignment.asset_id == asset_id,
        Assignment.status == AssignmentStatus.ACTIVE
    ).first()
    if active_assignment:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Cannot deactivate asset with active assignment"
        )

    asset.is_active = False

    # Audit log
    audit_entry = AuditLog(
        user_id=current_user.user_id,
        action="ASSET_DEACTIVATE",
        table_affected="assets",
        record_id=asset_id,
        details=f"Asset {asset.asset_name} deactivated",
    )
    db.add(audit_entry)
    db.commit()

    return {"message": "Asset deactivated successfully"}


@router.patch("/{asset_id}/reactivate")
def reactivate_asset(
    asset_id: str,
    db: Session = Depends(get_db),
    current_user=Depends(require_roles("Asset Manager", "System Administrator")),
):
    """
    Reactivate an asset. Sets is_active = True.
    """
    asset = db.query(Asset).filter(Asset.asset_id == asset_id).first()
    if not asset:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Asset with ID {asset_id} not found"
        )

    # Check if already active
    if asset.is_active:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Asset is already active"
        )

    asset.is_active = True

    # Audit log
    audit_entry = AuditLog(
        user_id=current_user.user_id,
        action="ASSET_REACTIVATE",
        table_affected="assets",
        record_id=asset_id,
        details=f"Asset {asset.asset_name} reactivated",
    )
    db.add(audit_entry)
    db.commit()

    return {"message": "Asset reactivated successfully"}


@router.patch("/{asset_id}/activate")
def activate_asset(
    asset_id: str,
    db: Session = Depends(get_db),
    current_user=Depends(require_roles("Asset Manager", "System Administrator")),
):
    """
    Activate a newly registered asset. Sets is_active = True.
    """
    asset = db.query(Asset).filter(Asset.asset_id == asset_id).first()
    if not asset:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Asset with ID {asset_id} not found"
        )

    # Check if already active
    if asset.is_active:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Asset is already active"
        )

    asset.is_active = True

    # Audit log
    audit_entry = AuditLog(
        user_id=current_user.user_id,
        action="ASSET_ACTIVATE",
        table_affected="assets",
        record_id=asset_id,
        details=f"Asset {asset.asset_name} activated",
    )
    db.add(audit_entry)
    db.commit()

    return {"message": "Asset activated successfully"}


@router.get("/export/pdf")
def export_assets_pdf(
    status: Optional[str] = Query(None),
    asset_type: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    _user=Depends(get_current_user),
):
    """
    Export assets to PDF. Respects the same filters as the list endpoint.
    Returns a downloadable PDF file.
    """
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from datetime import datetime
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="PDF export not available. Install reportlab."
        )

    # Query assets with filters
    q = db.query(Asset)
    if status:
        q = q.filter(Asset.status == status)
    if asset_type:
        q = q.filter(Asset.asset_type == asset_type)
    if search:
        q = q.filter(
            Asset.asset_name.ilike(f"%{search}%")
            | Asset.serial_number.ilike(f"%{search}%")
        )
    assets = q.order_by(Asset.created_at.desc()).all()

    # Create PDF with metadata
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=0.5*inch,
        leftMargin=0.5*inch,
        topMargin=0.75*inch,
        bottomMargin=0.5*inch,
        title="URSB Asset Management - Asset List",
        author="URSB Asset Management System",
        subject="Asset Export Report"
    )
    elements = []

    # Custom styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.darkblue,
        spaceAfter=20,
        alignment=1,  # Center
    )
    date_style = ParagraphStyle(
        'DateStyle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.gray,
        alignment=1,  # Center
    )

    # Title
    title = Paragraph("URSB Asset Management - Asset List", title_style)
    elements.append(title)

    # Date
    date_str = datetime.now().strftime("%B %d, %Y at %I:%M %p")
    date_para = Paragraph(f"Generated on: {date_str}", date_style)
    elements.append(date_para)
    elements.append(Spacer(1, 0.2*inch))

    # Table data with better spacing
    data = [["Asset ID", "Asset Name", "Type", "Serial Number", "Status", "Condition", "Cost (UGX)", "Department"]]
    for asset in assets:
        data.append([
            asset.asset_id,
            asset.asset_name,
            asset.asset_type.value if hasattr(asset.asset_type, "value") else str(asset.asset_type),
            asset.serial_number,
            asset.status.value if hasattr(asset.status, "value") else str(asset.status),
            asset.condition.value if hasattr(asset.condition, "value") else str(asset.condition),
            f"{float(asset.cost):,.2f}",
            asset.department or "—",
        ])

    # Create table with better column widths and word wrapping
    table = Table(data, colWidths=[1.2*inch, 2.0*inch, 1.0*inch, 1.2*inch, 1.0*inch, 1.0*inch, 1.0*inch, 1.0*inch], repeatRows=1)
    table.setStyle(TableStyle([
        # Header styling
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        
        # Data row styling
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('TOPPADDING', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        
        # Alternating row colors
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        
        # Grid
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        
        # Alignment for specific columns
        ('ALIGN', (6, 1), (6, -1), 'RIGHT'),  # Cost column right-aligned
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # Asset ID center-aligned
        ('ALIGN', (2, 1), (5, -1), 'CENTER'),  # Type, Status, Condition center-aligned
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=assets_export.pdf"}
    )


@router.get("/export/excel")
def export_assets_excel(
    status: Optional[str] = Query(None),
    asset_type: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    _user=Depends(get_current_user),
):
    """
    Export assets to Excel. Respects the same filters as the list endpoint.
    Returns a downloadable .xlsx file.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Excel export not available. Install openpyxl."
        )

    # Query assets with filters
    q = db.query(Asset)
    if status:
        q = q.filter(Asset.status == status)
    if asset_type:
        q = q.filter(Asset.asset_type == asset_type)
    if search:
        q = q.filter(
            Asset.asset_name.ilike(f"%{search}%")
            | Asset.serial_number.ilike(f"%{search}%")
        )
    assets = q.order_by(Asset.created_at.desc()).all()

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Assets"

    # Header row
    headers = ["Asset ID", "Name", "Type", "Category", "Serial", "Status", "Condition", "Cost", "Acquisition Date", "Supplier", "Department"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_num, asset in enumerate(assets, 2):
        ws.cell(row=row_num, column=1, value=asset.asset_id)
        ws.cell(row=row_num, column=2, value=asset.asset_name)
        ws.cell(row=row_num, column=3, value=asset.asset_type.value if hasattr(asset.asset_type, "value") else str(asset.asset_type))
        ws.cell(row=row_num, column=4, value=asset.category)
        ws.cell(row=row_num, column=5, value=asset.serial_number)
        ws.cell(row=row_num, column=6, value=asset.status.value if hasattr(asset.status, "value") else str(asset.status))
        ws.cell(row=row_num, column=7, value=asset.condition.value if hasattr(asset.condition, "value") else str(asset.condition))
        ws.cell(row=row_num, column=8, value=float(asset.cost))
        ws.cell(row=row_num, column=9, value=str(asset.acquisition_date))
        ws.cell(row=row_num, column=10, value=asset.supplier)
        ws.cell(row=row_num, column=11, value=asset.department or "")

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=assets_export.xlsx"}
    )
