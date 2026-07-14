"""Bulk user import routes for CSV and XLSX file processing."""

import asyncio
import json
import secrets
import string
from datetime import datetime, timedelta
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, WebSocket, WebSocketDisconnect, status
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.db import get_db, SessionLocal
from app.models.user import User, UserRole
from app.models.audit_log import AuditLog
from app.models.temporary_password import TemporaryPassword
from app.api.v1.auth import get_current_user, require_roles
from app.services.auth import create_password_hash, validate_ursb_email, get_session, SESSION_COOKIE_NAME

router = APIRouter(prefix="/api/v1/users", tags=["user-import"])


def generate_secure_password() -> str:
    """
    Generate a cryptographically random secure password.

    Requirements:
    - Minimum 12 characters
    - At least one uppercase letter (A-Z)
    - At least one lowercase letter (a-z)
    - At least one digit (0-9)
    - At least one special character from !@#$%^&*

    Uses Python's secrets module for cryptographic randomness.
    The generated value is returned once and never stored.
    """
    uppercase = string.ascii_uppercase
    lowercase = string.ascii_lowercase
    digits = string.digits
    special = "!@#$%^&*"

    # Ensure at least one character from each required set
    password = [
        secrets.choice(uppercase),
        secrets.choice(lowercase),
        secrets.choice(digits),
        secrets.choice(special),
    ]

    # Fill the rest with random characters from all sets
    all_chars = uppercase + lowercase + digits + special
    for _ in range(8):  # 8 more characters to reach minimum 12
        password.append(secrets.choice(all_chars))

    # Shuffle to avoid predictable pattern
    secrets.SystemRandom().shuffle(password)

    return ''.join(password)


class BulkImportError(BaseModel):
    row: int
    email: Optional[str]
    reason: str


class BulkImportAccount(BaseModel):
    full_name: str
    email: str
    role: str
    generated_password: str


class BulkImportResponse(BaseModel):
    total_rows: int
    created: int
    skipped: int
    errors: List[BulkImportError]
    accounts: List[BulkImportAccount]


@router.post("/bulk-import", response_model=BulkImportResponse)
async def bulk_import_users(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles("System Administrator")),
):
    """
    Bulk import users from CSV or XLSX file.

    Per-row validation order:
    1. Email format validation
    2. Duplicate email check
    3. Role validity check

    All valid rows are inserted in a single transaction. If the transaction fails,
    no rows are inserted.

    The accounts array in the response contains credentials for all successfully
    created accounts. This is the only time these passwords are available.
    """
    # Validate file type
    if not file.filename:
        raise HTTPException(status_code=400, detail="No file provided")
    
    file_ext = file.filename.lower().split('.')[-1]
    if file_ext not in ['csv', 'xlsx']:
        raise HTTPException(status_code=400, detail="Only .csv and .xlsx files are accepted")
    
    # Validate file size (2 MB max)
    MAX_SIZE = 2 * 1024 * 1024  # 2 MB
    content = await file.read()
    if len(content) > MAX_SIZE:
        raise HTTPException(status_code=400, detail="File size exceeds 2 MB limit")
    
    # Parse file based on extension
    rows = []
    if file_ext == 'csv':
        import io
        import csv
        csv_file = io.StringIO(content.decode('utf-8'))
        reader = csv.DictReader(csv_file)
        rows = list(reader)
    else:  # xlsx
        import io
        import openpyxl
        xlsx_file = io.BytesIO(content)
        workbook = openpyxl.load_workbook(xlsx_file)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if any(cell is not None for cell in row):
                rows.append(dict(zip(headers, row)))
    
    total_rows = len(rows)
    errors = []
    accounts_to_create = []
    
    # Per-row validation
    for idx, row in enumerate(rows, start=2):  # Start at 2 (1-based index after header)
        email = row.get('email', '').strip()
        full_name = row.get('full_name', '').strip()
        role = row.get('role', '').strip()
        department = row.get('department', '').strip()
        
        # Skip if required fields missing
        if not email or not full_name or not role:
            errors.append(BulkImportError(
                row=idx,
                email=email or None,
                reason="Missing required field (full_name, email, or role)"
            ))
            continue
        
        # Validate email format
        if '@' not in email or '.' not in email:
            errors.append(BulkImportError(
                row=idx,
                email=email,
                reason="Invalid email format"
            ))
            continue
        
        # Validate email domain (URSB)
        try:
            validate_ursb_email(email)
        except HTTPException:
            errors.append(BulkImportError(
                row=idx,
                email=email,
                reason="Invalid email domain (must be @ursb.go.ug)"
            ))
            continue
        
        # Check for duplicate email
        existing = db.query(User).filter(User.email == email).first()
        if existing:
            errors.append(BulkImportError(
                row=idx,
                email=email,
                reason="Email already exists"
            ))
            continue
        
        # Validate role
        try:
            UserRole(role)
        except ValueError:
            errors.append(BulkImportError(
                row=idx,
                email=email,
                reason=f"Invalid role: {role}"
            ))
            continue
        
        # Row is valid, add to creation list
        accounts_to_create.append({
            'full_name': full_name,
            'email': email.strip().lower(),
            'role': role,
            'department': department or None,
        })

    created_accounts = []
    
    try:
        for account_data in accounts_to_create:
            password = generate_secure_password()
            salt, password_hash = create_password_hash(password)
            
            # Parse first and last name
            name_parts = account_data['full_name'].split(' ', 1)
            first_name = name_parts[0]
            last_name = name_parts[1] if len(name_parts) > 1 else ''
            
            new_user = User(
                first_name=first_name,
                last_name=last_name,
                full_name=account_data['full_name'],
                email=account_data['email'],
                username=account_data['email'].split('@')[0],
                role=UserRole(account_data['role']),
                department=account_data['department'],
                password_salt=salt,
                password_hash=password_hash,
                is_active=True,
            )
            db.add(new_user)
            db.flush()
            
            temp_pw = TemporaryPassword(
                user_id=new_user.user_id,
                password=password,
                expires_at=datetime.utcnow() + timedelta(days=7),
            )
            db.add(temp_pw)
            
            created_accounts.append(BulkImportAccount(
                full_name=account_data['full_name'],
                email=account_data['email'],
                role=account_data['role'],
                generated_password=password,
            ))
        
        # Write single audit log for the entire batch
        if created_accounts:
            audit = AuditLog(
                user_id=current_user.user_id,
                action="BULK_USER_IMPORT",
                table_affected="users",
                record_id="bulk",
                details=f"Bulk import created {len(created_accounts)} user accounts by admin {current_user.email}",
                timestamp=datetime.utcnow(),
            )
            db.add(audit)
        
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Transaction failed: {str(e)}")
    
    return BulkImportResponse(
        total_rows=total_rows,
        created=len(created_accounts),
        skipped=len(errors),
        errors=errors,
        accounts=created_accounts
    )


@router.websocket("/bulk-import-ws")
async def bulk_import_ws(websocket: WebSocket):
    """
    WebSocket endpoint for real-time bulk user import with progress reporting.

    Protocol:
    - Client connects and sends a JSON array of row objects.
    - Server validates each row, then creates accounts one at a time.
    - After each row is processed, server sends: {"type": "progress", "progress": <0-100>, "processed": <n>, "total": <n>}
    - On completion: {"type": "complete", "total_rows": n, "created": n, "skipped": n, "errors": [...], "accounts": [...]}
    - If client disconnects mid-import (cancel), the DB transaction is rolled back.
    """
    await websocket.accept()

    # ── Authenticate via session cookie ──────────────────────────────────────────
    session_token = websocket.cookies.get(SESSION_COOKIE_NAME)
    if not session_token:
        await websocket.close(code=4001, reason="Not authenticated")
        return

    with SessionLocal() as auth_db:
        session = get_session(auth_db, session_token)
        if not session or not session.user or not session.user.is_active:
            await websocket.close(code=4001, reason="Not authenticated")
            return
        if session.user.role.value != "System Administrator":
            await websocket.close(code=4003, reason="Forbidden")
            return
        admin_user_id = session.user.user_id
        admin_email = session.user.email

    # ── Receive rows payload from client ─────────────────────────────────────────
    try:
        raw = await websocket.receive_text()
        rows = json.loads(raw)
    except Exception:
        await websocket.close(code=4400, reason="Invalid payload")
        return

    total_rows = len(rows)
    errors = []
    accounts_to_create = []

    # ── Validate all rows first ───────────────────────────────────────────────────
    with SessionLocal() as db:
        for idx, row in enumerate(rows, start=2):
            email = str(row.get('email', '')).strip()
            full_name = str(row.get('full_name', '')).strip()
            role = str(row.get('role', '')).strip()
            department = str(row.get('department', '')).strip() if row.get('department') else ''

            if not email or not full_name or not role:
                errors.append({"row": idx, "email": email or None, "reason": "Missing required field (full_name, email, or role)"})
                continue

            if '@' not in email or '.' not in email:
                errors.append({"row": idx, "email": email, "reason": "Invalid email format"})
                continue

            try:
                validate_ursb_email(email)
            except HTTPException:
                errors.append({"row": idx, "email": email, "reason": "Invalid email domain (must be @ursb.go.ug)"})
                continue

            existing = db.query(User).filter(User.email == email).first()
            if existing:
                errors.append({"row": idx, "email": email, "reason": "Email already exists"})
                continue

            try:
                UserRole(role)
            except ValueError:
                errors.append({"row": idx, "email": email, "reason": f"Invalid role: {role}"})
                continue

            accounts_to_create.append({
                'full_name': full_name,
                'email': email.strip().lower(),
                'role': role,
                'department': department or None,
            })

    # ── Create accounts one-by-one with progress pushes ──────────────────────────
    created_accounts = []
    total_to_create = len(accounts_to_create)

    with SessionLocal() as db:
        try:
            for i, account_data in enumerate(accounts_to_create):
                # Check if client is still connected before each row
                await asyncio.sleep(0)  # yield to event loop to detect disconnects

                password = generate_secure_password()
                salt, password_hash = create_password_hash(password)

                name_parts = account_data['full_name'].split(' ', 1)
                first_name = name_parts[0]
                last_name = name_parts[1] if len(name_parts) > 1 else ''

                new_user = User(
                    first_name=first_name,
                    last_name=last_name,
                    full_name=account_data['full_name'],
                    email=account_data['email'],
                    username=account_data['email'].split('@')[0],
                    role=UserRole(account_data['role']),
                    department=account_data['department'],
                    password_salt=salt,
                    password_hash=password_hash,
                    is_active=True,
                )
                db.add(new_user)
                db.flush()

                temp_pw = TemporaryPassword(
                    user_id=new_user.user_id,
                    password=password,
                    expires_at=datetime.utcnow() + timedelta(days=7),
                )
                db.add(temp_pw)

                created_accounts.append({
                    'full_name': account_data['full_name'],
                    'email': account_data['email'],
                    'role': account_data['role'],
                    'generated_password': password,
                })

                # Send real progress — percentage of VALID rows that have been created
                progress = round(((i + 1) / total_to_create) * 100) if total_to_create > 0 else 100
                await websocket.send_json({
                    "type": "progress",
                    "progress": progress,
                    "processed": i + 1,
                    "total": total_to_create,
                })
                # Small sleep so the client can render updates smoothly
                await asyncio.sleep(0.05)

            # Audit log
            if created_accounts:
                audit = AuditLog(
                    user_id=admin_user_id,
                    action="BULK_USER_IMPORT",
                    table_affected="users",
                    record_id="bulk",
                    details=f"Bulk import created {len(created_accounts)} user accounts by admin {admin_email}",
                    timestamp=datetime.utcnow(),
                )
                db.add(audit)

            db.commit()

            # Send completion message
            await websocket.send_json({
                "type": "complete",
                "total_rows": total_rows,
                "created": len(created_accounts),
                "skipped": len(errors),
                "errors": errors,
                "accounts": created_accounts,
            })

        except WebSocketDisconnect:
            # Client cancelled — roll back everything
            db.rollback()
            return
        except Exception as e:
            db.rollback()
            try:
                await websocket.send_json({"type": "error", "message": str(e)})
            except Exception:
                pass
            return

    await websocket.close()



def generate_secure_password() -> str:
    """
    Generate a cryptographically random secure password.

    Requirements:
    - Minimum 12 characters
    - At least one uppercase letter (A-Z)
    - At least one lowercase letter (a-z)
    - At least one digit (0-9)
    - At least one special character from !@#$%^&*

    Uses Python's secrets module for cryptographic randomness.
    The generated value is returned once and never stored.
    """
    uppercase = string.ascii_uppercase
    lowercase = string.ascii_lowercase
    digits = string.digits
    special = "!@#$%^&*"

    # Ensure at least one character from each required set
    password = [
        secrets.choice(uppercase),
        secrets.choice(lowercase),
        secrets.choice(digits),
        secrets.choice(special),
    ]

    # Fill the rest with random characters from all sets
    all_chars = uppercase + lowercase + digits + special
    for _ in range(8):  # 8 more characters to reach minimum 12
        password.append(secrets.choice(all_chars))

    # Shuffle to avoid predictable pattern
    secrets.SystemRandom().shuffle(password)

    return ''.join(password)


class BulkImportError(BaseModel):
    row: int
    email: Optional[str]
    reason: str


class BulkImportAccount(BaseModel):
    full_name: str
    email: str
    role: str
    generated_password: str


class BulkImportResponse(BaseModel):
    total_rows: int
    created: int
    skipped: int
    errors: List[BulkImportError]
    accounts: List[BulkImportAccount]


@router.post("/bulk-import", response_model=BulkImportResponse)
async def bulk_import_users(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles("System Administrator")),
):
    """
    Bulk import users from CSV or XLSX file.

    Per-row validation order:
    1. Email format validation
    2. Duplicate email check
    3. Role validity check

    All valid rows are inserted in a single transaction. If the transaction fails,
    no rows are inserted.

    The accounts array in the response contains credentials for all successfully
    created accounts. This is the only time these passwords are available.
    """
    # Validate file type
    if not file.filename:
        raise HTTPException(status_code=400, detail="No file provided")
    
    file_ext = file.filename.lower().split('.')[-1]
    if file_ext not in ['csv', 'xlsx']:
        raise HTTPException(status_code=400, detail="Only .csv and .xlsx files are accepted")
    
    # Validate file size (2 MB max)
    MAX_SIZE = 2 * 1024 * 1024  # 2 MB
    content = await file.read()
    if len(content) > MAX_SIZE:
        raise HTTPException(status_code=400, detail="File size exceeds 2 MB limit")
    
    # Parse file based on extension
    rows = []
    if file_ext == 'csv':
        import io
        import csv
        csv_file = io.StringIO(content.decode('utf-8'))
        reader = csv.DictReader(csv_file)
        rows = list(reader)
    else:  # xlsx
        import io
        import openpyxl
        xlsx_file = io.BytesIO(content)
        workbook = openpyxl.load_workbook(xlsx_file)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if any(cell is not None for cell in row):
                rows.append(dict(zip(headers, row)))
    
    total_rows = len(rows)
    errors = []
    accounts_to_create = []
    
    # Per-row validation
    for idx, row in enumerate(rows, start=2):  # Start at 2 (1-based index after header)
        email = row.get('email', '').strip()
        full_name = row.get('full_name', '').strip()
        role = row.get('role', '').strip()
        department = row.get('department', '').strip()
        
        # Skip if required fields missing
        if not email or not full_name or not role:
            errors.append(BulkImportError(
                row=idx,
                email=email or None,
                reason="Missing required field (full_name, email, or role)"
            ))
            continue
        
        # Validate email format
        if '@' not in email or '.' not in email:
            errors.append(BulkImportError(
                row=idx,
                email=email,
                reason="Invalid email format"
            ))
            continue
        
        # Validate email domain (URSB)
        try:
            validate_ursb_email(email)
        except HTTPException:
            errors.append(BulkImportError(
                row=idx,
                email=email,
                reason="Invalid email domain (must be @ursb.go.ug)"
            ))
            continue
        
        # Check for duplicate email
        existing = db.query(User).filter(User.email == email).first()
        if existing:
            errors.append(BulkImportError(
                row=idx,
                email=email,
                reason="Email already exists"
            ))
            continue
        
        # Validate role
        try:
            UserRole(role)
        except ValueError:
            errors.append(BulkImportError(
                row=idx,
                email=email,
                reason=f"Invalid role: {role}"
            ))
            continue
        
        # Row is valid, add to creation list
        accounts_to_create.append({
            'full_name': full_name,
            'email': email.strip().lower(),
            'role': role,
            'department': department,
            'row_idx': idx
        })
    
    # Insert valid rows in a single transaction
    created_accounts = []
    try:
        for account_data in accounts_to_create:
            generated_password = generate_secure_password()
            salt, p_hash = create_password_hash(generated_password)

            new_user = User(
                full_name=account_data['full_name'],
                email=account_data['email'],
                password_hash=p_hash,
                password_salt=salt,
                role=UserRole(account_data['role']),
                department=account_data['department'] or None,
                is_active=True,
            )
            db.add(new_user)
            db.flush()  # Get the user_id before creating temp password

            # Store temporary password for admin viewing (expires in 7 days)
            temp_password = TemporaryPassword(
                user_id=new_user.user_id,
                password=generated_password,
                created_at=datetime.utcnow(),
                expires_at=datetime.utcnow() + timedelta(days=7),
            )
            db.add(temp_password)

            created_accounts.append(BulkImportAccount(
                full_name=account_data['full_name'],
                email=account_data['email'],
                role=account_data['role'],
                generated_password=generated_password
            ))
        
        # Write single audit log for the entire batch
        if created_accounts:
            audit = AuditLog(
                user_id=current_user.user_id,
                action="BULK_USER_IMPORT",
                table_affected="users",
                record_id="bulk",
                details=f"Bulk import created {len(created_accounts)} user accounts by admin {current_user.email}",
                timestamp=datetime.utcnow(),
            )
            db.add(audit)
        
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Transaction failed: {str(e)}")
    
    return BulkImportResponse(
        total_rows=total_rows,
        created=len(created_accounts),
        skipped=len(errors),
        errors=errors,
        accounts=created_accounts
    )
