"""Asset routes: list, filter, search, and register assets."""

import io
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.db import get_db
from app.models.asset import Asset, AssetCondition, AssetStatus, AssetType, SourceType
from app.models.assignment import Assignment, AssignmentStatus
from app.models.audit_log import AuditLog
from app.api.v1.auth import get_current_user, require_roles
from app.services import asset_service

router = APIRouter(prefix="/api/v1/assets", tags=["assets"])


# ── Response Schemas ────────────────────────────────────────────────────────────────

class AssetDetailResponse(BaseModel):
    """Response for GET /api/v1/assets/{id} - full asset with all related data."""
    asset_id: str
    asset_name: str
    asset_type: str
    category: str
    serial_number: str
    condition: str
    status: str
    source_type: str
    procurement_ref: str | None
    cost: float
    acquisition_date: str
    supplier: str
    department: str | None
    current_custodian_id: str | None
    is_active: bool
    created_at: str
    updated_at: str
    current_custodian: dict | None
    assignment_history: List[dict]
    maintenance_history: List[dict]
    transfer_history: List[dict]
    disposal_record: dict | None

    class Config:
        from_attributes = True


class AssetOut(BaseModel):
    asset_id: str
    asset_name: str
    asset_type: str
    category: str
    serial_number: str
    condition: str
    status: str
    cost: float
    acquisition_date: str
    supplier: str
    department: str | None
    created_at: str

    class Config:
        from_attributes = True


class AssetCreate(BaseModel):
    name: str
    asset_type: str
    serial_number: str
    condition: str
    cost: float
    department: Optional[str] = None
    acquisition_date: str
    status: str
    category: str
    supplier: str
    source_type: str


class AssetUpdateRequest(BaseModel):
    """Request body for PUT /api/v1/assets/{id}. All fields optional."""
    asset_name: Optional[str] = None
    category: Optional[str] = None
    condition: Optional[str] = None
    status: Optional[str] = None
    department: Optional[str] = None
    current_custodian_id: Optional[int] = None
    supplier: Optional[str] = None
    procurement_ref: Optional[str] = None
    cost: Optional[float] = None


# ── Helpers ─────────────────────────────────────────────────────────────────────────

def _serialize_asset(a: Asset) -> AssetOut:
    return AssetOut(
        asset_id=a.asset_id,
        asset_name=a.asset_name,
        asset_type=a.asset_type.value if hasattr(a.asset_type, "value") else str(a.asset_type),
        category=a.category,
        serial_number=a.serial_number,
        condition=a.condition.value if hasattr(a.condition, "value") else str(a.condition),
        status=a.status.value if hasattr(a.status, "value") else str(a.status),
        cost=float(a.cost),
        acquisition_date=str(a.acquisition_date),
        supplier=a.supplier,
        department=a.department,
        created_at=str(a.created_at),
    )


def _build_detail_response(asset: Asset) -> AssetDetailResponse:
    current_custodian = None
    if asset.current_custodian:
        current_custodian = {
            "id": str(asset.current_custodian.id),
            "first_name": asset.current_custodian.first_name,
            "last_name": asset.current_custodian.last_name,
            "email": asset.current_custodian.email,
            "department": asset.current_custodian.department,
        }

    assignment_history = [
        {
            "assignment_id": a.assignment_id,
            "assigned_to_name": a.assigned_to_user.full_name if a.assigned_to_user else None,
            "assigned_by_name": a.assigned_by_user.full_name if a.assigned_by_user else None,
            "assignment_date": str(a.assignment_date),
            "return_date": str(a.return_date) if a.return_date else None,
            "status": a.status.value if hasattr(a.status, "value") else str(a.status),
            "notes": a.notes,
        }
        for a in sorted(asset.assignments, key=lambda x: x.assignment_date, reverse=True)
    ]

    maintenance_history = [
        {
            "maintenance_id": m.maintenance_id,
            "service_date": str(m.service_date),
            "service_provider": m.service_provider,
            "description": m.description,
            "cost": float(m.cost),
            "next_service_date": str(m.next_service_date) if m.next_service_date else None,
        }
        for m in sorted(asset.maintenance_records, key=lambda x: x.service_date, reverse=True)
    ]

    transfer_history = [
        {
            "transfer_id": t.transfer_id,
            "from_user_name": t.from_user.full_name if t.from_user else None,
            "to_user_name": t.to_user.full_name if t.to_user else None,
            "transfer_date": str(t.transfer_date),
            "reason": t.reason,
            "acknowledged_at": str(t.acknowledged_at) if t.acknowledged_at else None,
        }
        for t in sorted(asset.transfers, key=lambda x: x.transfer_date, reverse=True)
    ]

    disposal_record = None
    if asset.disposal_records:
        d = asset.disposal_records[0]
        disposal_record = {
            "disposal_date": str(d.disposal_date),
            "disposal_method": d.disposal_method.value if hasattr(d.disposal_method, "value") else str(d.disposal_method),
            "reason": d.reason,
            "authorised_by_name": d.authorised_by_user.full_name if d.authorised_by_user else None,
        }

    return AssetDetailResponse(
        asset_id=asset.asset_id,
        asset_name=asset.asset_name,
        asset_type=asset.asset_type.value if hasattr(asset.asset_type, "value") else str(asset.asset_type),
        category=asset.category,
        serial_number=asset.serial_number,
        condition=asset.condition.value if hasattr(asset.condition, "value") else str(asset.condition),
        status=asset.status.value if hasattr(asset.status, "value") else str(asset.status),
        source_type=asset.source_type.value if hasattr(asset.source_type, "value") else str(asset.source_type),
        procurement_ref=asset.procurement_ref,
        cost=float(asset.cost),
        acquisition_date=str(asset.acquisition_date),
        supplier=asset.supplier,
        department=asset.department,
        current_custodian_id=str(asset.current_custodian_id) if asset.current_custodian_id else None,
        is_active=asset.is_active,
        created_at=str(asset.created_at),
        updated_at=str(asset.updated_at),
        current_custodian=current_custodian,
        assignment_history=assignment_history,
        maintenance_history=maintenance_history,
        transfer_history=transfer_history,
        disposal_record=disposal_record,
    )


# ── Endpoints ────────────────────────────────────────────────────────────────────────

@router.post("", response_model=AssetOut, status_code=status.HTTP_201_CREATED)
def create_asset(
    body: AssetCreate,
    db: Session = Depends(get_db),
    current_user=Depends(require_roles("Asset Manager")),
):
    """Register a new asset. Asset Manager only. SRS AM-R01."""
    asset = asset_service.create_asset(db, body, current_user.user_id)
    return _serialize_asset(asset)


@router.get("", response_model=List[AssetOut])
def list_assets(
    status: Optional[str] = Query(None),
    asset_type: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    _user=Depends(get_current_user),
):
    """List assets with optional filters. All authenticated roles. SRS AM-P03."""
    assets, _ = asset_service.list_assets(db, status=status, asset_type=asset_type, search=search)
    return [_serialize_asset(a) for a in assets]


@router.get("/export/csv")
def export_assets_csv(
    status: Optional[str] = Query(None),
    asset_type: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    department: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    _user=Depends(get_current_user),
):
    """Export assets to CSV. All authenticated roles. SRS AM-P04."""
    csv_content = asset_service.export_assets_csv(db, {
        "status": status,
        "asset_type": asset_type,
        "search": search,
        "department": department,
    })
    return StreamingResponse(
        iter([csv_content]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=assets_export.csv"},
    )


@router.get("/{asset_id}", response_model=AssetDetailResponse)
def get_asset_detail(
    asset_id: str,
    db: Session = Depends(get_db),
    _user=Depends(get_current_user),
):
    """Fetch a single asset with full detail. All authenticated roles. SRS AM-P05."""
    asset = asset_service.get_asset(db, asset_id)
    return _build_detail_response(asset)


@router.put("/{asset_id}", response_model=AssetDetailResponse)
def update_asset(
    asset_id: str,
    body: AssetUpdateRequest,
    db: Session = Depends(get_db),
    current_user=Depends(require_roles("Asset Manager", "System Administrator")),
):
    """Update an asset. Asset Manager and System Administrator only. SRS AM-R03."""
    asset = asset_service.update_asset(db, asset_id, body, current_user.user_id)
    return _build_detail_response(asset)


@router.patch("/{asset_id}/deactivate")
def deactivate_asset(
    asset_id: str,
    db: Session = Depends(get_db),
    current_user=Depends(require_roles("Asset Manager", "System Administrator")),
):
    """Deactivate an asset. Asset Manager and System Administrator only. SRS AM-R05."""
    asset = asset_service.get_asset(db, asset_id)

    if not asset.is_active:
        raise HTTPException(400, detail="Asset is already inactive")
    if asset.status == AssetStatus.DISPOSED:
        raise HTTPException(400, detail="Cannot deactivate a disposed asset")

    active_assignment = (
        db.query(Assignment)
        .filter(Assignment.asset_id == asset_id, Assignment.status == AssignmentStatus.ACTIVE)
        .first()
    )
    if active_assignment:
        raise HTTPException(400, detail="Cannot deactivate asset with active assignment")

    asset.is_active = False
    db.add(AuditLog(
        user_id=current_user.user_id,
        action="ASSET_DEACTIVATE",
        table_affected="assets",
        record_id=asset_id,
        details=f"Asset {asset.asset_name} deactivated",
    ))
    db.commit()
    return {"message": "Asset deactivated successfully"}


@router.patch("/{asset_id}/reactivate")
def reactivate_asset(
    asset_id: str,
    db: Session = Depends(get_db),
    current_user=Depends(require_roles("Asset Manager", "System Administrator")),
):
    """Reactivate an asset. Asset Manager and System Administrator only. SRS AM-R05."""
    asset = asset_service.get_asset(db, asset_id)
    if asset.is_active:
        raise HTTPException(400, detail="Asset is already active")
    asset.is_active = True
    db.add(AuditLog(
        user_id=current_user.user_id,
        action="ASSET_REACTIVATE",
        table_affected="assets",
        record_id=asset_id,
        details=f"Asset {asset.asset_name} reactivated",
    ))
    db.commit()
    return {"message": "Asset reactivated successfully"}


@router.patch("/{asset_id}/activate")
def activate_asset(
    asset_id: str,
    db: Session = Depends(get_db),
    current_user=Depends(require_roles("Asset Manager", "System Administrator")),
):
    """Activate a newly registered asset. Asset Manager and System Administrator only."""
    asset = asset_service.get_asset(db, asset_id)
    if asset.is_active:
        raise HTTPException(400, detail="Asset is already active")
    asset.is_active = True
    db.add(AuditLog(
        user_id=current_user.user_id,
        action="ASSET_ACTIVATE",
        table_affected="assets",
        record_id=asset_id,
        details=f"Asset {asset.asset_name} activated",
    ))
    db.commit()
    return {"message": "Asset activated successfully"}


@router.get("/export/pdf")
def export_assets_pdf(
    status: Optional[str] = Query(None),
    asset_type: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    _user=Depends(get_current_user),
):
    """Export assets to PDF. All authenticated roles. SRS AM-P06."""
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from datetime import datetime
    except ImportError:
        raise HTTPException(500, detail="PDF export not available. Install reportlab.")

    assets, _ = asset_service.list_assets(db, status=status, asset_type=asset_type, search=search)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=0.5*inch, leftMargin=0.5*inch, topMargin=0.75*inch, bottomMargin=0.5*inch)
    elements = []
    styles = getSampleStyleSheet()

    elements.append(Paragraph("URSB Asset Management - Asset List", styles["Heading1"]))
    elements.append(Paragraph(f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}", styles["Normal"]))
    elements.append(Spacer(1, 0.2*inch))

    data = [["Asset ID", "Asset Name", "Type", "Serial Number", "Status", "Condition", "Cost (UGX)", "Department"]]
    for a in assets:
        data.append([
            a.asset_id, a.asset_name,
            a.asset_type.value if hasattr(a.asset_type, "value") else str(a.asset_type),
            a.serial_number,
            a.status.value if hasattr(a.status, "value") else str(a.status),
            a.condition.value if hasattr(a.condition, "value") else str(a.condition),
            f"{float(a.cost):,.2f}", a.department or "—",
        ])

    table = Table(data, colWidths=[1.2*inch, 2.0*inch, 1.0*inch, 1.2*inch, 1.0*inch, 1.0*inch, 1.0*inch, 1.0*inch], repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)

    return StreamingResponse(buffer, media_type="application/pdf", headers={"Content-Disposition": "attachment; filename=assets_export.pdf"})


@router.get("/export/excel")
def export_assets_excel(
    status: Optional[str] = Query(None),
    asset_type: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    _user=Depends(get_current_user),
):
    """Export assets to Excel. All authenticated roles. SRS AM-P04."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        raise HTTPException(500, detail="Excel export not available. Install openpyxl.")

    assets, _ = asset_service.list_assets(db, status=status, asset_type=asset_type, search=search)

    wb = Workbook()
    ws = wb.active
    ws.title = "Assets"
    headers = ["Asset ID", "Name", "Type", "Category", "Serial", "Status", "Condition", "Cost", "Acquisition Date", "Supplier", "Department"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    for row_num, a in enumerate(assets, 2):
        ws.cell(row=row_num, column=1, value=a.asset_id)
        ws.cell(row=row_num, column=2, value=a.asset_name)
        ws.cell(row=row_num, column=3, value=a.asset_type.value if hasattr(a.asset_type, "value") else str(a.asset_type))
        ws.cell(row=row_num, column=4, value=a.category)
        ws.cell(row=row_num, column=5, value=a.serial_number)
        ws.cell(row=row_num, column=6, value=a.status.value if hasattr(a.status, "value") else str(a.status))
        ws.cell(row=row_num, column=7, value=a.condition.value if hasattr(a.condition, "value") else str(a.condition))
        ws.cell(row=row_num, column=8, value=float(a.cost))
        ws.cell(row=row_num, column=9, value=str(a.acquisition_date))
        ws.cell(row=row_num, column=10, value=a.supplier)
        ws.cell(row=row_num, column=11, value=a.department or "")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=assets_export.xlsx"})